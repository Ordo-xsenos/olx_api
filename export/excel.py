from datetime import datetime
from io import BytesIO
from typing import Dict, List

from openpyxl import Workbook
from openpyxl.utils import get_column_letter


HEADERS = [
    ("category", "Категория"),
    ("title", "Название"),
    ("price", "Цена"),
    ("currency", "Валюта"),
    ("location", "Локация"),
    ("precise_location", "Точная локация"),
    ("url", "Ссылка"),
    ("created_at", "Дата добавления"),
]


def build_excel(rows: List[Dict]) -> BytesIO:
    output = BytesIO()
    wb = Workbook()
    ws = wb.active
    ws.title = "OLX"

    ws.append([label for _, label in HEADERS])

    for row in rows:
        values = []
        for key, _ in HEADERS:
            value = row.get(key)
            if isinstance(value, datetime) and value.tzinfo is not None:
                value = value.replace(tzinfo=None)
            values.append(value)
        ws.append(values)

    for idx, (key, label) in enumerate(HEADERS, start=1):
        max_len = len(str(label))
        for cell in ws[get_column_letter(idx)]:
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[get_column_letter(idx)].width = min(max_len + 2, 60)

    wb.save(output)
    output.seek(0)
    return output
